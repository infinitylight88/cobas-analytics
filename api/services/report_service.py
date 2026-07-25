import io
from datetime import date

import openpyxl
from sqlalchemy import func
from sqlalchemy.orm import Session

from api.database.models import ArchiveFile, Calibration, PatientResult, QCResult


class ReportService:

    def daily(self, db: Session, target_date: str):
        d = date.fromisoformat(target_date)
        samples = (
            db.query(func.count(func.distinct(PatientResult.sample_id)))
            .filter(func.date(PatientResult.run_datetime) == d)
            .scalar()
        )
        patients = (
            db.query(func.count(func.distinct(PatientResult.patient_identifier)))
            .filter(func.date(PatientResult.run_datetime) == d)
            .scalar()
        )
        tests = (
            db.query(func.count(PatientResult.result_id))
            .filter(func.date(PatientResult.run_datetime) == d)
            .scalar()
        )
        qc_runs = (
            db.query(func.count(QCResult.qc_id))
            .filter(func.date(QCResult.run_datetime) == d)
            .scalar()
        )
        cals = (
            db.query(func.count(Calibration.calibration_id))
            .filter(func.date(Calibration.run_datetime) == d)
            .scalar()
        )
        analyte_rows = (
            db.query(PatientResult.analyte_code, func.count(PatientResult.result_id).label("n"))
            .filter(func.date(PatientResult.run_datetime) == d)
            .group_by(PatientResult.analyte_code)
            .order_by(func.count(PatientResult.result_id).desc())
            .all()
        )
        return {
            "date": target_date,
            "unique_samples": samples,
            "unique_patients": patients,
            "total_tests": tests,
            "qc_runs": qc_runs,
            "calibrations": cals,
            "analyte_breakdown": [{"analyte_code": r.analyte_code, "count": r.n} for r in analyte_rows],
        }

    def monthly(self, db: Session, year: int, month: int):
        rows = (
            db.query(
                func.date(PatientResult.run_datetime).label("date"),
                func.count(func.distinct(PatientResult.sample_id)).label("samples"),
                func.count(func.distinct(PatientResult.patient_identifier)).label("patients"),
                func.count(PatientResult.result_id).label("tests"),
            )
            .filter(
                func.extract("year", PatientResult.run_datetime) == year,
                func.extract("month", PatientResult.run_datetime) == month,
            )
            .group_by(func.date(PatientResult.run_datetime))
            .order_by(func.date(PatientResult.run_datetime))
            .all()
        )
        total_samples = sum(r.samples for r in rows)
        total_patients = db.query(func.count(func.distinct(PatientResult.patient_identifier))).filter(
            func.extract("year", PatientResult.run_datetime) == year,
            func.extract("month", PatientResult.run_datetime) == month,
        ).scalar()
        total_tests = sum(r.tests for r in rows)
        return {
            "year": year,
            "month": month,
            "working_days": len(rows),
            "total_samples": total_samples,
            "total_patients": total_patients,
            "total_tests": total_tests,
            "daily_breakdown": [
                {"date": str(r.date), "samples": r.samples, "patients": r.patients, "tests": r.tests}
                for r in rows
            ],
        }

    def qc_report(self, db: Session):
        summary = (
            db.query(
                QCResult.analyte_code,
                QCResult.control_name,
                func.count(QCResult.qc_id).label("runs"),
                func.avg(QCResult.measured_value).label("mean"),
                func.min(QCResult.run_datetime).label("first"),
                func.max(QCResult.run_datetime).label("last"),
            )
            .group_by(QCResult.analyte_code, QCResult.control_name)
            .order_by(QCResult.analyte_code, QCResult.control_name)
            .all()
        )
        return {
            "total_qc_results": db.query(func.count(QCResult.qc_id)).scalar(),
            "analytes_with_qc": len({r.analyte_code for r in summary}),
            "controls_used": len({r.control_name for r in summary}),
            "summary": [
                {
                    "analyte_code": r.analyte_code,
                    "control_name": r.control_name,
                    "runs": r.runs,
                    "mean_measured": round(float(r.mean), 4) if r.mean else None,
                    "first_run": str(r.first),
                    "last_run": str(r.last),
                }
                for r in summary
            ],
        }

    def calibration_report(self, db: Session):
        summary = (
            db.query(
                Calibration.analyte_code,
                func.count(Calibration.calibration_id).label("calibrations"),
                func.max(Calibration.run_datetime).label("last_cal"),
            )
            .group_by(Calibration.analyte_code)
            .order_by(Calibration.analyte_code)
            .all()
        )
        return {
            "total_calibrations": db.query(func.count(Calibration.calibration_id)).scalar(),
            "analytes_calibrated": len(summary),
            "summary": [
                {"analyte_code": r.analyte_code, "calibrations": r.calibrations, "last_calibration": str(r.last_cal)}
                for r in summary
            ],
        }

    def actg_report(self, db: Session):
        rows = (
            db.query(
                func.date(PatientResult.run_datetime).label("date"),
                func.count(func.distinct(PatientResult.patient_identifier)).label("patients"),
                func.count(PatientResult.result_id).label("results"),
            )
            .filter(PatientResult.patient_program == "ACTG")
            .group_by(func.date(PatientResult.run_datetime))
            .order_by(func.date(PatientResult.run_datetime))
            .all()
        )
        return {
            "total_patients": db.query(func.count(func.distinct(PatientResult.patient_identifier))).filter(PatientResult.patient_program == "ACTG").scalar(),
            "total_results": db.query(func.count(PatientResult.result_id)).filter(PatientResult.patient_program == "ACTG").scalar(),
            "daily": [{"date": str(r.date), "patients": r.patients, "results": r.results} for r in rows],
        }

    def workload_report(self, db: Session):
        rows = (
            db.query(
                func.to_char(PatientResult.run_datetime, "YYYY-MM").label("month"),
                func.count(func.distinct(PatientResult.sample_id)).label("samples"),
                func.count(func.distinct(PatientResult.patient_identifier)).label("patients"),
                func.count(PatientResult.result_id).label("tests"),
            )
            .group_by(func.to_char(PatientResult.run_datetime, "YYYY-MM"))
            .order_by(func.to_char(PatientResult.run_datetime, "YYYY-MM"))
            .all()
        )
        return {
            "monthly_workload": [
                {"month": r.month, "unique_samples": r.samples, "unique_patients": r.patients, "total_tests": r.tests}
                for r in rows
            ]
        }

    def export_excel(self, db: Session):
        wb = openpyxl.Workbook()

        # Sheet 1: Patient Results Summary
        ws1 = wb.active
        ws1.title = "Patient Results"
        rows = db.query(PatientResult).order_by(PatientResult.run_datetime.desc()).all()
        ws1.append(["Result ID", "Run DateTime", "Patient ID", "Patient Initials", "Sample ID",
                    "Analyte Code", "Analyte Name", "Result Value", "Units",
                    "Reference Flag", "Status", "Program"])
        for r in rows:
            ws1.append([r.result_id, str(r.run_datetime), r.patient_identifier, r.patient_initials,
                        r.sample_id, r.analyte_code, r.analyte_name,
                        float(r.result_value) if r.result_value else None,
                        r.units, r.reference_flag, r.status, r.patient_program])

        # Sheet 2: QC Results
        ws2 = wb.create_sheet("QC Results")
        qc_rows = db.query(QCResult).order_by(QCResult.run_datetime.desc()).all()
        ws2.append(["QC ID", "Run DateTime", "Analyte", "Control Name", "Control Lot",
                    "Assigned Value", "Measured Value", "SD", "Units"])
        for r in qc_rows:
            ws2.append([r.qc_id, str(r.run_datetime), r.analyte_code, r.control_name, r.control_lot,
                        float(r.assigned_value) if r.assigned_value else None,
                        float(r.measured_value) if r.measured_value else None,
                        float(r.sd) if r.sd else None, r.units])

        # Sheet 3: Calibrations
        ws3 = wb.create_sheet("Calibrations")
        cal_rows = db.query(Calibration).order_by(Calibration.run_datetime.desc()).all()
        ws3.append(["Cal ID", "Run DateTime", "Analyte", "Calibration Type", "Reagent Lot", "Slope", "Offset"])
        for r in cal_rows:
            ws3.append([r.calibration_id, str(r.run_datetime), r.analyte_code, r.calibration_type,
                        r.reagent_lot,
                        float(r.slope) if r.slope else None,
                        float(r.offset) if r.offset else None])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
